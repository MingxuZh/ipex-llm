import os
import os
import xlwt 
from xlwt import Workbook 
import re
wb = Workbook() 
row=0
col=0
sheet1 = wb.add_sheet('Sheet 1')
sheet1.write(row,col,'kernel')
sheet1.write(row,col+1,'mult')
sheet1.write(row,col+2,'time')
sheet1.write(row,col+3,'c_freq')
sheet1.write(row,col+4,'uc_freq')
row=row+1
col=col
# Getting the current work directory (cwd)
thisdir = os.getcwd()
str="perf,cpu"
kernel=[]
time=[]
mult=[]
mult_get=[]
kernel_str=[]
i=0
thr_list_each_file=[]


import openpyxl
from openpyxl import load_workbook

def Average(lst):
    return sum(lst) / len(lst)

# Manual Settings
reference_file=load_workbook("/home/guojian/amx_bf16/AMX_INT8.xlsx")
bdnn_logs_path="/workspace/logs/TMULPVP_RUN2/benchdnn_20221125004024"
path_emon="/home/guojian/amx_bf16/emon/"
final_result = "/home/guojian/amx_bf16/amx_bf16_E9I.0093.D22.xlsx"

for r, d, f in os.walk(bdnn_logs_path):
    for file in f:
        text_file = open(os.path.join(r, file), "r")
        lines = text_file.readlines()
        ws = reference_file['Sheet']
        file_name_list=file.split("_")
        print(file_name_list)
        counter=0
        for col in ws['A']:
            counter=counter+1
            print(file_name_list[7])
            if file_name_list[7] in col.value:
               print("matched")
               break
        print(counter)

        val=[]
        for line in lines:
            if str in line:
                print(line)
                i=i+1
                 
                thr_list=line.split(',')
                ind=thr_list.index("cpu")
                val.append(float(thr_list[ind+8]))
                print(float(thr_list[ind+8]))

        cellref=ws.cell(counter, 3)
        cellref.value=Average(val);
        print("Average",cellref.value)
 
print("i",i)
row=1
col=4
#emon path that save rsync file
'''
rsync -ratlz dcso@dcsometrics.intel.com:/data2/metrics/emondata/data/leirong/localhost.localdomain/TMULPVP_RUN2/BDNN_AMX_INT8*/BDNN_AMX_INT8_*/BDNN_AMX_INT8__ww39_sh_spr2*.xlsx .  --progress
intel321
'''
from openpyxl import load_workbook
for filename in os.listdir(path_emon):
    file_name_list=[]
    if filename.endswith(".xlsx"):
       file_name_list=filename.split("_r")
       print(file_name_list)
       ind=int(re.search(r'\d+', file_name_list[1]).group())
       print(ind)
       
       wb_emon = load_workbook(path_emon+filename)
       print(path_emon+filename)
       sheet = wb_emon["system view"]
       x2=sheet.cell(row=2, column=1)
       print(x2.value)
       x3 = sheet.cell(row=2, column=2)
       print(x3.value)
       cellref=ws.cell(ind+2, 4)
       cellref.value=float(x3.value)
       x4 = sheet.cell(row=64, column=2)
       print(x4.value)
       cellref=ws.cell(ind+2, 5)
       cellref.value=float(x4.value)

#result file
print("="*30+"end"+"="*30)
reference_file.save("{}".format(final_result))
